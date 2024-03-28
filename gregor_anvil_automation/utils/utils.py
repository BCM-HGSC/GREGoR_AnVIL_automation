from pathlib import Path
import csv

import addict
import yaml
from openpyxl import Workbook, load_workbook


from .exceptions import InputPathDoesNotExistError
from .types import Sample


def get_table_samples(input_path: Path) -> dict[str, list[Sample]]:
    """Get tables from either an excel path or directory filled with TSVs"""
    if not input_path.exists():
        raise InputPathDoesNotExistError(input_path)
    if ".xlsx" in input_path.suffixes:
        return get_table_samples_by_excel(input_path)
    if input_path.is_dir():
        return get_table_samples_by_directory(input_path)
    raise NotImplementedError


def get_table_samples_by_directory(dir_path: Path) -> dict[str, list[Sample]]:
    """Gets every TSV file in the directory."""
    data = {}
    for file in dir_path.glob("*"):
        if "xlsx" in file.suffix:
            data[file.stem] = get_table_samples_by_excel(file)
        if "tsv" in file.suffix:
            print(file)
            data[file.stem] = parse_file(file, "\t")
    return data


def get_table_samples_by_excel(input_file: Path) -> dict[str, list[Sample]]:
    """Reads the given excel file path and gets the samples"""
    workbook: Workbook = load_workbook(input_file)
    sheet = workbook.active
    max_column = sheet.max_column
    samples = []
    headers = []
    for i in range(1, max_column + 1):
        if header := sheet.cell(row=1, column=i).value:
            headers.append(header.strip().lower().replace(" ", "_"))
    for row_cells in sheet.iter_rows(min_row=2):
        if all(cell.value is None or cell.value.strip() == "" for cell in row_cells):
            continue
        sample = {
            header: (
                ""
                if row_cells[idx].value is None
                else str(row_cells[idx].value).strip()
            )
            for idx, header in enumerate(headers)
        }
        sample["row_number"] = row_cells[0].row
        samples.append(sample)
    return samples


def parse_yaml(yaml_path: Path) -> addict.Dict:
    """Parses a yaml file and return Iterator"""
    with open(yaml_path, encoding="utf-8") as fin:
        return addict.Dict(yaml.safe_load(fin.read()))


def parse_file(file_path: Path, delimiter: str) -> addict.Dict:
    """Parses a file"""
    data = []
    with open(file_path, "r", encoding="utf-8") as fin:
        reader = csv.DictReader(fin, delimiter=delimiter)
        for idx, line in enumerate(reader, 2):
            line["row_number"] = idx
            data.append(line)
    return data


def generate_file(
    file_path: Path, data_headers: list[str], data: list[dict[str, str]], delimiter: str
):
    """Generates either a csv or tsv file depending on the passed in delimiter"""
    with open(file_path, "w", encoding="utf-8") as file:
        writer = csv.DictWriter(
            f=file, fieldnames=data_headers, delimiter=delimiter, extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(data)
